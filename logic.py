"""
Ops Pulse -- reads the staging spreadsheet's 'Orders' + 'Not Shipped'
tabs (the SAME staging sheet sync_orders.py's daily job and the orders_status_native
app both write into), builds one unified per-order dataset, and computes Fulfillment /
Delivery KPIs for a chosen date range -- either one period on its own, or two periods
side by side for an "are we improving or not" comparison, plus a rule-based Summary
that calls out what got better, what got worse, and which metrics currently sit far
enough from target to count as a weak point.

Sep 2026, per Mahmoud: reads Orders + Not Shipped TOGETHER on purpose. Orders alone
undercounts Total/Cancelled/Pending Orders -- an order only gets a row in a raw
tracking sheet (and so in the staging Orders tab) once it's physically WITH the
shipping company; anything Cancelled or still Pending before that point only ever shows
up in Not Shipped (see orders_status_native/README.md and consolidation_tool's Orders
Status Check page for the full story this app is built on top of).

Pure logic, no Streamlit and no gspread CALLS made from here beyond load_orders_data()
accepting an already-authorized client -- everything else takes plain DataFrames, so it
can be unit-tested with synthetic data (same philosophy as clean.py / engine.py
elsewhere in this project family).
"""
import io
import re
import datetime as dt

import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Read-only tool -- only needs read scopes, unlike sync_orders.py / orders_status_native
# which also write to these sheets.
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets.readonly',
    'https://www.googleapis.com/auth/drive.readonly',
]

# Same real staging sheet baked into orders_status_native/logic.py (Aug 2026, confirmed)
# -- this tool reads the SAME staging spreadsheet, just never writes to it. app.py's
# staging_spreadsheet_id secret, if set, overrides this.
STAGING_SPREADSHEET_ID_DEFAULT = '1dZMqtqvnxe6GspH0C10AvXECB74NP-ZjDG_BihMOkmg'


def get_client(service_account_info):
    creds = Credentials.from_service_account_info(service_account_info, scopes=SCOPES)
    return gspread.authorize(creds)


ORDERS_TAB = 'Orders'
NOT_SHIPPED_TAB = 'Not Shipped'

# Both tabs share this same 13-column shape at minimum (orders_status_native's own
# ORDERS_HEADER). The staging 'Orders' tab additionally carries 'Delivery Date' and
# 'Customer Phone' at the end (added Sep 2026, sync_orders.py) which 'Not Shipped' does
# not have (a Shopify "all orders" export has no shipping-company data yet) -- those 2
# just come out blank/NaT for Not Shipped rows in the unified frame, not an error.
CORE_HEADER = [
    'Source', 'Reference Number', 'Shipping Date', 'Order Date', 'Country', 'City',
    'Order Value', 'Status', 'New Customer Orders', 'Returning Customer Orders',
    'Needs Review', 'Review Reason', 'Last Synced At (UTC)',
]
EXTRA_ORDERS_COLUMNS = ['Delivery Date', 'Customer Phone']
FULL_HEADER = CORE_HEADER + EXTRA_ORDERS_COLUMNS

STATUSES = ['Delivered', 'Returned', 'Pending', 'Cancelled']

GOOGLE_SHEETS_EPOCH = dt.date(1899, 12, 30)
_ISO_DATE_RE = re.compile(r'^\s*(\d{4})-(\d{2})-(\d{2})\s*$')

# Default "how many days late counts as not-on-time" -- fallback only, used if a
# country shows up with no entry in a per-market target dict at all. The real per-market
# numbers now live in DEFAULT_DELIVERY_WINDOWS below (Sep 2026, CEO Q3 2026 scorecard --
# "GC KPIs, Younes" -- Delivery Time KPI row), plugged in via app.py's sidebar.
DEFAULT_ON_TIME_TARGET_DAYS = 5

# Per-market transit windows in days, Shipping Date -> Delivery Date (Sep 2026, from the
# CEO Q3 2026 scorecard's "Delivery Time" KPI row: "Within window: UAE 2-3d, KSA 5-7d,
# QA 5-7d, KW 10-12d, OM (via UAE) 7-10d"). (low, high) per market -- 'high' doubles as
# the On-Time Delivery Rate's per-market target (an order delivered within the window,
# i.e. gap <= high, counts as on-time; see _resolve_on_time_target()). 'low' is only
# used for the Delivery Time KPI's own Below/Target/Exceed band (see
# classify_delivery_time_band()). Iraq has no window in the scorecard -- Delivery Date
# isn't captured there yet (see the module docstring) -- so it's left out on purpose
# rather than guessed; Iraq is excluded from On-Time Delivery Rate and from Delivery
# Time bands until a real window exists for it, exactly like it already is from the
# Delivery Time average itself.
DEFAULT_DELIVERY_WINDOWS = {
    'UAE': (2, 3),
    'OM': (7, 10),
    'SA': (5, 7),
    'QA': (5, 7),
    'KW': (10, 12),
}

# Below/Target/Exceed cutoffs for the 2 rate-based KPIs that the scorecard gives ONE
# flat number for regardless of market (unlike Delivery Time's per-market windows
# above). 'below' isn't stored -- classify_band() treats "under target" as Below by
# definition, so only the 2 real cutoffs need a home.
DEFAULT_KPI_BANDS = {
    'on_time_rate': {'target': 0.95, 'exceed': 0.97},        # scorecard: <90% Below
    'net_delivery_rate': {'target': 0.92, 'exceed': 0.95},   # scorecard: <88% Below
}

# Net Delivery Rate's "matured cohort" gate (Sep 2026, CEO scorecard: "no rate quoted
# while >10% in transit"): the share of a period's Shipped orders still Pending
# (in transit -- shipped but not yet Delivered/Returned) has to be at or under this
# share before the rate is quoted at all -- see _metrics_for_slice()'s net_delivery_rate
# / net_delivery_matured / in_transit_share.
DEFAULT_NET_DELIVERY_MATURED_THRESHOLD = 0.10

# Weak-point thresholds (Sep 2026, per Mahmoud: "sensible defaults I define, but keep
# them adjustable from the UI too" -- see app.py's threshold sliders). All in
# "how much WORSE than the baseline period counts as a weak point" terms. Rates are in
# percentage points (e.g. 2.0 = 2pp), time metrics are in days.
DEFAULT_WEAK_POINT_THRESHOLDS = {
    'delivered_rate_pp': 2.0,           # dropped by >= 2pp
    'cancelled_rate_pp': 2.0,           # rose by >= 2pp
    'returned_rate_pp': 2.0,            # rose by >= 2pp
    'pending_rate_pp': 3.0,             # rose by >= 3pp
    'fulfillment_lead_time_days': 1.0,  # got slower by >= 1 day
    'delivery_time_days': 1.0,          # got slower by >= 1 day
    'on_time_rate_pp': 5.0,             # dropped by >= 5pp
    'net_delivery_rate_pp': 5.0,        # dropped by >= 5pp
}

METRIC_LABELS = {
    'delivered_rate': 'Delivered rate',
    'returned_rate': 'Returned rate',
    'cancelled_rate': 'Cancelled rate',
    'pending_rate': 'Pending rate',
    'fulfillment_lead_time_days': 'Avg. fulfillment lead time (days)',
    'delivery_time_days': 'Avg. delivery time (days)',
    'on_time_rate': 'On-time delivery rate',
    'net_delivery_rate': 'Net delivery rate',
    'new_rate': 'New-customer rate',
    'returning_rate': 'Returning-customer rate',
}

# direction=+1 means "higher is better" (e.g. delivered_rate, on_time_rate);
# direction=-1 means "lower is better" (e.g. cancelled_rate, lead time).
METRIC_DIRECTION = {
    'delivered_rate': 1,
    'returned_rate': -1,
    'cancelled_rate': -1,
    'pending_rate': -1,
    'fulfillment_lead_time_days': -1,
    'delivery_time_days': -1,
    'on_time_rate': 1,
    'net_delivery_rate': 1,
    'new_rate': 1,
    'returning_rate': 1,
}


def classify_band(value, target, exceed):
    """value: a 0-1 rate (e.g. on_time_rate, net_delivery_rate). target/exceed: the 2
    cutoffs from DEFAULT_KPI_BANDS (or the sidebar's adjusted versions) -- 'target'
    means "at or above the Target line but short of Exceed" (the scorecard's Green
    band: hitting the number exactly still counts as on target, not short of it).
    Returns 'exceed' / 'target' / 'below', or None if value or target is missing."""
    if value is None or target is None:
        return None
    if exceed is not None and value >= exceed:
        return 'exceed'
    if value >= target:
        return 'target'
    return 'below'


BAND_LABELS = {'below': 'Below target', 'target': 'On target', 'exceed': 'Exceeding target'}


def per_country_kpi_rows(metrics, delivery_windows, on_time_bands, net_delivery_bands):
    """One row per market combining Net Delivery Rate / On-Time Delivery Rate /
    Delivery Time with their CEO-scorecard Below/Target/Exceed bands. THE single source
    of truth for this breakdown -- both app.py's on-screen "Logistics KPIs by market"
    table AND the xlsx export's own "Logistics KPIs" sheet call this same function (Sep
    2026, per Mahmoud: the bands shown on screen need to actually be in the download
    too, not just the raw numbers -- computing them in one place instead of twice is
    what keeps the two views from ever silently drifting apart).
    metrics: a compute_period_metrics() return value (has 'per_country').
    delivery_windows: {country: (low, high)} (see DEFAULT_DELIVERY_WINDOWS).
    on_time_bands / net_delivery_bands: (target, exceed) tuples (see DEFAULT_KPI_BANDS).
    """
    ot_target, ot_exceed = on_time_bands
    nd_target, nd_exceed = net_delivery_bands
    rows = []
    for country in sorted(metrics.get('per_country', {})):
        m = metrics['per_country'][country]
        window = delivery_windows.get(country)
        matured = m.get('net_delivery_matured')
        rows.append({
            'country': country,
            'net_delivery_rate': m.get('net_delivery_rate') if matured else None,
            'net_delivery_matured': matured,
            'in_transit_share': m.get('in_transit_share'),
            'shipped_n': m.get('shipped_n'),
            'net_delivery_band': classify_band(m.get('net_delivery_rate'), nd_target, nd_exceed) if matured else None,
            'on_time_rate': m.get('on_time_rate'),
            'on_time_band': classify_band(m.get('on_time_rate'), ot_target, ot_exceed),
            'delivery_time_days': m.get('delivery_time_days'),
            'delivery_window': window,
            'delivery_time_band': classify_delivery_time_band(m.get('delivery_time_days'), window),
        })
    return rows


def classify_delivery_time_band(avg_days, window):
    """window: (low, high) from DEFAULT_DELIVERY_WINDOWS (or the sidebar's adjusted
    version) for ONE market -- Delivery Time is the only KPI here where 'Target' is a
    range, not a single cutoff, so it needs its own classifier. Below (Red) = averaged
    above the window; Target (Green) = inside the window; Exceed (Stretch) = at or below
    the window's lower bound -- mirrors the scorecard's 3 bands for this KPI exactly.
    Returns None (no band) if there's no window for this market (Iraq) or no data."""
    if avg_days is None or not window:
        return None
    lo, hi = window
    if avg_days <= lo:
        return 'exceed'
    if avg_days <= hi:
        return 'target'
    return 'below'


# ---------------------------------------------------------------------------
# Reading the 2 staging tabs into one unified DataFrame
# ---------------------------------------------------------------------------

def _cell_to_date(value):
    """Both staging tabs write dates as plain ISO text ('YYYY-MM-DD') -- see
    sync_orders.py / orders_status_native's own fmt_date(). Handles that directly, plus
    a genuine Sheets date-serial number as a safety net for a cell someone later
    hand-edited into a native Sheets date type (which would come back as a serial
    number under UNFORMATTED_VALUE, not the original ISO text). Returns pd.Timestamp
    (or pd.NaT) rather than datetime.date, so date-arithmetic/.dt accessors work
    directly on a column built from this without an extra conversion step."""
    if value in (None, ''):
        return pd.NaT
    if isinstance(value, bool):
        return pd.NaT
    if isinstance(value, (int, float)):
        if 0 < value < 100000:
            return pd.Timestamp(GOOGLE_SHEETS_EPOCH) + pd.Timedelta(days=value)
        return pd.NaT
    s = str(value).strip()
    if not s:
        return pd.NaT
    m = _ISO_DATE_RE.match(s)
    if m:
        try:
            return pd.Timestamp(dt.date(int(m.group(1)), int(m.group(2)), int(m.group(3))))
        except ValueError:
            return pd.NaT
    try:
        return pd.Timestamp(pd.to_datetime(s))
    except Exception:
        return pd.NaT


def _cell_to_number(value):
    if value in (None, ''):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(',', '').strip())
    except (TypeError, ValueError):
        return None


def sheet_values_to_df(values, origin_tab):
    """values: raw 2D list from gspread's get_values(value_render_option=
    'UNFORMATTED_VALUE') for either tab, header row included (or []/[[]] for an empty/
    missing tab). Returns a normalized DataFrame with a fixed column set regardless of
    which tab it came from -- Not Shipped rows get NaT/None for the 2 Orders-only
    columns rather than erroring."""
    cols = FULL_HEADER + ['_origin_tab']
    if not values or len(values) < 2:
        return pd.DataFrame(columns=cols)

    header = [str(h).strip() for h in values[0]]
    body = values[1:]
    # Pad/truncate every row to the header's width -- a raw sheet read can have short
    # trailing rows (Sheets omits fully-blank trailing cells).
    width = len(header)
    body = [list(r) + [''] * (width - len(r)) if len(r) < width else r[:width] for r in body]
    raw = pd.DataFrame(body, columns=header)

    # Reindex to the full expected column set -- tolerant of a tab missing a column
    # entirely (Not Shipped has no Delivery Date/Customer Phone at all) rather than
    # raising, and tolerant of duplicate/blank header cells by just taking the first
    # match for each expected name.
    df = pd.DataFrame(index=raw.index)
    for col in FULL_HEADER:
        df[col] = raw[col] if col in raw.columns else None

    out = pd.DataFrame(index=df.index)
    out['_ref_key'] = df['Reference Number'].astype(str).str.strip().str.upper().str.lstrip('#')
    out['_source'] = df['Source'].astype(str).str.strip()
    out['_order_date'] = df['Order Date'].map(_cell_to_date)
    out['_shipping_date'] = df['Shipping Date'].map(_cell_to_date)
    out['_delivery_date'] = df['Delivery Date'].map(_cell_to_date)
    out['_country'] = df['Country'].astype(str).str.strip().str.upper()
    out['_city'] = df['City'].astype(str).str.strip()
    out['_order_value'] = df['Order Value'].map(_cell_to_number)
    out['_status'] = df['Status'].astype(str).str.strip()
    out['_new_customer'] = df['New Customer Orders'].map(_cell_to_number).fillna(0)
    out['_returning_customer'] = df['Returning Customer Orders'].map(_cell_to_number).fillna(0)
    out['_customer_phone'] = df['Customer Phone'].astype(str).str.strip()
    out.loc[out['_customer_phone'].isin(['None', 'nan']), '_customer_phone'] = ''
    out['_origin_tab'] = origin_tab
    # Drop rows with no reference number at all (fully-blank trailing rows some sheets
    # carry) -- nothing meaningful to report on without a join key.
    out = out[out['_ref_key'] != '']
    return out.reset_index(drop=True)


def load_orders_data(gc, staging_spreadsheet_id):
    """Reads both tabs live and returns one combined DataFrame, de-duplicated by
    Reference Number: an order present in BOTH tabs (the day it ships, before Not
    Shipped has been cleaned up -- see orders_status_native's "Clean up Not Shipped"
    feature) keeps only its Orders-tab row, since that one carries the real
    shipping-company data (Delivery Date, Customer Phone, the real Status) and Not
    Shipped's copy of it is stale by definition."""
    sh = gc.open_by_key(staging_spreadsheet_id)

    def _read_tab(tab_name):
        try:
            ws = sh.worksheet(tab_name)
        except Exception:
            return sheet_values_to_df([], tab_name)
        values = ws.get_values(value_render_option='UNFORMATTED_VALUE')
        return sheet_values_to_df(values, tab_name)

    orders_df = _read_tab(ORDERS_TAB)
    not_shipped_df = _read_tab(NOT_SHIPPED_TAB)
    return combine_tabs(orders_df, not_shipped_df)


def combine_tabs(orders_df, not_shipped_df):
    """Split out from load_orders_data() so it can be unit-tested without a live
    Google Sheets connection -- see the module docstring."""
    combined = pd.concat([orders_df, not_shipped_df], ignore_index=True)
    if combined.empty:
        return combined
    combined['_priority'] = (combined['_origin_tab'] == ORDERS_TAB).astype(int)
    combined = combined.sort_values('_priority', ascending=False, kind='stable')
    combined = combined.drop_duplicates(subset='_ref_key', keep='first')
    combined = combined.drop(columns=['_priority']).reset_index(drop=True)
    return combined


# ---------------------------------------------------------------------------
# Period metrics
# ---------------------------------------------------------------------------

def filter_period(df, start_date, end_date, countries=None):
    """start_date/end_date: date-like (datetime.date or pd.Timestamp), inclusive.
    Filters on Order Date -- the one date field present and populated on every row
    regardless of which tab it came from (Shipping/Delivery Date are blank for most
    Pending/Cancelled rows, which would silently drop them from a shipping-date-based
    filter and understate exactly the statuses this tool exists to surface)."""
    if df.empty:
        return df
    start_ts = pd.Timestamp(start_date)
    end_ts = pd.Timestamp(end_date)
    mask = df['_order_date'].notna() & (df['_order_date'] >= start_ts) & (df['_order_date'] <= end_ts)
    if countries:
        wanted = {c.upper() for c in countries}
        mask &= df['_country'].isin(wanted)
    return df[mask].copy()


def _avg_day_gap(sub, from_col, to_col, extra_mask=None):
    """Average (to_col - from_col) in days, only over rows where both are present and
    to_col is not before from_col (a negative gap means bad data -- a typo'd date, a
    hand-edited cell -- not a real same-day-or-earlier delivery; excluded rather than
    letting it drag the average down artificially)."""
    m = sub[from_col].notna() & sub[to_col].notna()
    if extra_mask is not None:
        m &= extra_mask
    pairs = sub[m]
    if pairs.empty:
        return None, 0
    gap_days = (pairs[to_col] - pairs[from_col]).dt.days
    gap_days = gap_days[gap_days >= 0]
    if gap_days.empty:
        return None, 0
    return float(gap_days.mean()), int(len(gap_days))


def _resolve_on_time_target(country, on_time_target_days):
    if on_time_target_days is None:
        return None
    if isinstance(on_time_target_days, dict):
        return on_time_target_days.get(country, on_time_target_days.get('_default'))
    return on_time_target_days  # single number applied to every market


def _metrics_for_slice(sub, on_time_target_days=None, net_delivery_matured_threshold=DEFAULT_NET_DELIVERY_MATURED_THRESHOLD):
    """sub: an already-filtered (period + country) DataFrame slice. Returns the flat
    metrics dict shared by both the overall and the per-country breakdown."""
    total = len(sub)
    status_counts = {s: int((sub['_status'] == s).sum()) for s in STATUSES}
    status_rates = {s: (status_counts[s] / total if total else None) for s in STATUSES}
    # Sum of Order Value per status (Sep 2026, per Mahmoud: "عاوز احسب فلوس الاوردرات
    # الوصلت و الرجعت و الاتكنسل والبندج" -- the money behind each status, not just how
    # many orders). skipna=True so a blank Order Value cell doesn't turn a real sum
    # into NaN for the whole status.
    status_value = {s: float(sub.loc[sub['_status'] == s, '_order_value'].sum(skipna=True)) for s in STATUSES}

    fulfillment_days, fulfillment_n = _avg_day_gap(sub, '_order_date', '_shipping_date')

    delivered = sub[sub['_status'] == 'Delivered']
    # Delivery Time / On-Time Delivery Rate both measure the SHIPPING leg only --
    # Shipping Date -> Delivery Date, NOT Order Date -> Delivery Date (Sep 2026,
    # reconciled against the CEO Q3 2026 scorecard: the scorecard's own windows are
    # "fulfilment to delivery" / "system transit windows", i.e. the shipping company's
    # own transit time, extracted from their system -- it does not include however long
    # an order sat before it even reached them. That earlier stretch is a DIFFERENT
    # metric already tracked on its own: Fulfillment lead time (Order Date ->
    # Shipping Date, just above). Mixing the two into one Order-Date-based gap would
    # blame the shipping company's KPI for delays that actually happened in
    # confirmation/prep -- a Call Center/Ops problem, not a delivery one -- and would
    # make every per-market window too tight to mean anything.
    delivery_days, delivery_n = _avg_day_gap(delivered, '_shipping_date', '_delivery_date')

    on_time_rate, on_time_n = None, 0
    if on_time_target_days is not None:
        have_both = delivered[delivered['_shipping_date'].notna() & delivered['_delivery_date'].notna()]
        if not have_both.empty:
            gap = (have_both['_delivery_date'] - have_both['_shipping_date']).dt.days
            targets = have_both['_country'].map(lambda c: _resolve_on_time_target(c, on_time_target_days))
            evaluable = targets.notna()
            if evaluable.any():
                on_time = (gap[evaluable] <= targets[evaluable])
                on_time_rate = float(on_time.mean())
                on_time_n = int(evaluable.sum())

    new_n = int(sub['_new_customer'].sum())
    returning_n = int(sub['_returning_customer'].sum())
    new_rate = new_n / total if total else None
    returning_rate = returning_n / total if total else None

    total_value = float(sub['_order_value'].sum(skipna=True)) if total else 0.0
    avg_value = (total_value / total) if total else None

    # Net Delivery Rate = (Delivered - Returned) / Shipped, "matured cohorts only" (Sep
    # 2026, CEO Q3 2026 scorecard). "Shipped" = orders that actually reached the
    # shipping company -- i.e. have a Shipping Date at all -- which naturally excludes
    # Cancelled (cancelled before ever shipping, so it never gets one -- see
    # orders_status_native/logic.py) and any Pending order still sitting in Not Shipped
    # waiting to go out. A Pending order WITH a Shipping Date is a different thing: it
    # HAS shipped, just hasn't resolved (delivered or returned) yet -- "still in
    # transit". "Matured" is a COHORT-level gate, not a per-order age check: if more
    # than net_delivery_matured_threshold's share of this period's Shipped orders are
    # still in that in-transit state, the whole period's rate is withheld (None) rather
    # than quoted on an incomplete outcome -- see in_transit_share / net_delivery_matured.
    shipped_mask = sub['_shipping_date'].notna()
    shipped_n = int(shipped_mask.sum())
    shipped_status_counts = {s: int((sub.loc[shipped_mask, '_status'] == s).sum()) for s in STATUSES}
    in_transit_share = (shipped_status_counts['Pending'] / shipped_n) if shipped_n else None
    net_delivery_matured = (in_transit_share is not None and in_transit_share <= net_delivery_matured_threshold)
    net_delivery_rate = (
        (shipped_status_counts['Delivered'] - shipped_status_counts['Returned']) / shipped_n
    ) if net_delivery_matured else None

    return {
        'total_orders': total,
        'status_counts': status_counts,
        'status_value': status_value,
        'status_rates': status_rates,
        'delivered_rate': status_rates['Delivered'],
        'returned_rate': status_rates['Returned'],
        'pending_rate': status_rates['Pending'],
        'cancelled_rate': status_rates['Cancelled'],
        'fulfillment_lead_time_days': fulfillment_days,
        'fulfillment_lead_time_n': fulfillment_n,
        'delivery_time_days': delivery_days,
        'delivery_time_n': delivery_n,
        'on_time_rate': on_time_rate,
        'on_time_n': on_time_n,
        'shipped_n': shipped_n,
        'in_transit_share': in_transit_share,
        'net_delivery_matured': net_delivery_matured,
        'net_delivery_rate': net_delivery_rate,
        'new_rate': new_rate,
        'returning_rate': returning_rate,
        'total_order_value': total_value,
        'avg_order_value': avg_value,
    }


def compute_period_metrics(df, start_date, end_date, countries=None, on_time_target_days=None,
                            net_delivery_matured_threshold=DEFAULT_NET_DELIVERY_MATURED_THRESHOLD):
    """Top-level entry point for ONE period. Returns the overall metrics dict (see
    _metrics_for_slice) plus a 'per_country' dict of the same shape keyed by country
    code, and the raw filtered row count for a sanity check in the UI."""
    sub = filter_period(df, start_date, end_date, countries)
    overall = _metrics_for_slice(sub, on_time_target_days, net_delivery_matured_threshold)
    per_country = {}
    if not sub.empty:
        for country, grp in sub.groupby('_country'):
            per_country[country] = _metrics_for_slice(grp, on_time_target_days, net_delivery_matured_threshold)
    overall['per_country'] = per_country
    overall['start_date'] = pd.Timestamp(start_date).date()
    overall['end_date'] = pd.Timestamp(end_date).date()
    return overall


# ---------------------------------------------------------------------------
# Two-period comparison + rule-based Summary
# ---------------------------------------------------------------------------

def compare_periods(df, period_a, period_b, countries=None, on_time_target_days=None,
                     net_delivery_matured_threshold=DEFAULT_NET_DELIVERY_MATURED_THRESHOLD):
    """period_a / period_b: (start_date, end_date) tuples. 'a' is the baseline/earlier
    period, 'b' is what it's being compared against -- deltas are always b - a, so a
    positive delta on a "higher is better" metric (e.g. delivered_rate) means period b
    improved on period a."""
    metrics_a = compute_period_metrics(df, *period_a, countries=countries, on_time_target_days=on_time_target_days,
                                        net_delivery_matured_threshold=net_delivery_matured_threshold)
    metrics_b = compute_period_metrics(df, *period_b, countries=countries, on_time_target_days=on_time_target_days,
                                        net_delivery_matured_threshold=net_delivery_matured_threshold)

    def _delta(key, a_dict, b_dict, as_pp=False):
        va, vb = a_dict.get(key), b_dict.get(key)
        if va is None or vb is None:
            return None
        d = vb - va
        return d * 100 if as_pp else d

    rate_keys = ['delivered_rate', 'returned_rate', 'pending_rate', 'cancelled_rate', 'on_time_rate',
                 'net_delivery_rate', 'new_rate', 'returning_rate']
    day_keys = ['fulfillment_lead_time_days', 'delivery_time_days']

    def _deltas_for(a_dict, b_dict):
        d = {k: _delta(k, a_dict, b_dict, as_pp=True) for k in rate_keys}
        d.update({k: _delta(k, a_dict, b_dict, as_pp=False) for k in day_keys})
        d['total_orders'] = (b_dict['total_orders'] - a_dict['total_orders']) if (a_dict['total_orders'] is not None and b_dict['total_orders'] is not None) else None
        d['total_order_value'] = b_dict['total_order_value'] - a_dict['total_order_value']
        # Per-status count/value deltas (Sep 2026, per Mahmoud: the money behind
        # Delivered/Returned/Cancelled/Pending, as its own metric card like the rest --
        # not just a number in a table) -- keyed 'status_count_<Status>' /
        # 'status_value_<Status>' so app.py's metric cards can pull a delta the same
        # way every other card does.
        for s in STATUSES:
            d[f'status_count_{s}'] = b_dict['status_counts'][s] - a_dict['status_counts'][s]
            d[f'status_value_{s}'] = b_dict['status_value'][s] - a_dict['status_value'][s]
        return d

    overall_deltas = _deltas_for(metrics_a, metrics_b)
    per_country_deltas = {}
    all_countries = set(metrics_a['per_country']) | set(metrics_b['per_country'])
    for c in all_countries:
        a_c = metrics_a['per_country'].get(c)
        b_c = metrics_b['per_country'].get(c)
        if a_c is None or b_c is None:
            continue  # country only present in one period -- nothing to compare yet
        per_country_deltas[c] = _deltas_for(a_c, b_c)

    return {
        'period_a': metrics_a,
        'period_b': metrics_b,
        'deltas': overall_deltas,
        'per_country_deltas': per_country_deltas,
    }


def _fmt_pp(value):
    return f"{abs(value):.1f}pp"


def _fmt_days(value):
    return f"{abs(value):.1f}d"


def generate_summary(comparison, thresholds=None):
    """Rule-based, not AI-guessed: walks the comparison's deltas against `thresholds`
    (see DEFAULT_WEAK_POINT_THRESHOLDS) and buckets each evaluable metric into 'good'
    (moved enough in the right direction), 'weak_points' (moved enough in the WRONG
    direction -- these are what change per Mahmoud's "weak points I need to fix" ask),
    or left out of both if the move was inside the threshold (noise, not a real trend).
    Runs once for the overall numbers and once per country, so a market-specific
    problem hiding inside an OK overall number still surfaces."""
    th = dict(DEFAULT_WEAK_POINT_THRESHOLDS)
    if thresholds:
        th.update(thresholds)

    def _classify(scope, deltas):
        good, bad = [], []
        rate_metric_to_threshold = {
            'delivered_rate': 'delivered_rate_pp',
            'cancelled_rate': 'cancelled_rate_pp',
            'returned_rate': 'returned_rate_pp',
            'pending_rate': 'pending_rate_pp',
            'on_time_rate': 'on_time_rate_pp',
            'net_delivery_rate': 'net_delivery_rate_pp',
        }
        for metric, threshold_key in rate_metric_to_threshold.items():
            d = deltas.get(metric)
            if d is None:
                continue
            threshold = th.get(threshold_key)
            if threshold is None:
                continue
            direction = METRIC_DIRECTION[metric]
            signed = d * direction  # positive signed == moved in the GOOD direction
            word = 'rose' if d > 0 else 'dropped'
            entry = {'scope': scope, 'metric': metric, 'delta_pp': d,
                     'message': f"{METRIC_LABELS[metric]} {word} {_fmt_pp(d)} in {scope}."}
            if signed >= threshold:
                good.append(entry)
            elif signed <= -threshold:
                bad.append(entry)

        for metric, threshold_key in [('fulfillment_lead_time_days', 'fulfillment_lead_time_days'),
                                       ('delivery_time_days', 'delivery_time_days')]:
            d = deltas.get(metric)
            if d is None:
                continue
            threshold = th.get(threshold_key)
            if threshold is None:
                continue
            direction = METRIC_DIRECTION[metric]
            signed = d * direction
            word = 'got slower by' if d > 0 else 'got faster by'
            entry = {'scope': scope, 'metric': metric, 'delta_days': d,
                     'message': f"{METRIC_LABELS[metric]} {word} {_fmt_days(d)} in {scope}."}
            if signed >= threshold:
                good.append(entry)
            elif signed <= -threshold:
                bad.append(entry)
        return good, bad

    good_overall, bad_overall = _classify('overall', comparison['deltas'])
    good_countries, bad_countries = [], []
    for country, deltas in comparison['per_country_deltas'].items():
        g, b = _classify(country, deltas)
        good_countries.extend(g)
        bad_countries.extend(b)

    # Weak points: sorted worst-first so the most urgent issue leads the Summary tab.
    def _severity(item):
        return abs(item.get('delta_pp', item.get('delta_days', 0)))

    weak_points = sorted(bad_overall + bad_countries, key=_severity, reverse=True)
    good_points = sorted(good_overall + good_countries, key=_severity, reverse=True)

    return {
        'good': good_points,
        'weak_points': weak_points,
        'thresholds_used': th,
    }


# ---------------------------------------------------------------------------
# Excel export (Sep 2026, per Mahmoud: "عاوز اقدر اسحب الكلام ده اكسيل بكل
# التفاصيل و الشارتس" -- export whatever's currently on screen, full detail
# tables + charts, as one .xlsx). Charts here are NATIVE Excel chart objects
# (openpyxl.chart), built off the same tables written into the sheet -- not
# picture/image exports of the on-screen Plotly charts. That keeps them
# editable/native once opened in Excel and avoids adding an image-rendering
# dependency (kaleido) that isn't otherwise needed anywhere in this app or
# guaranteed to work on Streamlit Cloud.
#
# Deliberately NO per-bar data labels on any chart here (tried once, reverted -- Sep
# 2026, per Mahmoud). openpyxl writes correct, minimal dLbls XML (verified directly:
# showVal=1 with every other show* flag explicitly 0), but Google Sheets' xlsx-chart
# importer doesn't honour those individual flags -- it renders every label as
# "<series>, <category>, <value>" regardless, which is unreadable once several bars
# share a chart. Real Excel/LibreOffice would have rendered the clean version fine, but
# Mahmoud opens these in Google Sheets, so the numbers stay in the table right next to
# each chart instead -- the axis (country/status) and legend (period, on 2-series
# charts) are the only in-chart labels, and both render identically everywhere.
# ---------------------------------------------------------------------------

_XLSX_HEADER_FILL = PatternFill(start_color='1F4E3D', end_color='1F4E3D', fill_type='solid')
_XLSX_HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF')
_XLSX_BODY_FONT = Font(name='Arial')
_XLSX_TITLE_FONT = Font(name='Arial', bold=True, size=13)
_XLSX_SUBTITLE_FONT = Font(name='Arial', italic=True, color='555555')

# Light background fills for Below/Target/Exceed band cells (Sep 2026) -- same
# red/green/purple language as the Streamlit app's 🔴/🟢/🟣 captions, just as a cell
# fill instead of an emoji (Excel/Google Sheets render emoji fine too, but a fill reads
# at a glance across a whole table the way the on-screen captions do one card at a time).
_BAND_FILLS = {
    'below': PatternFill(start_color='FBE1E1', end_color='FBE1E1', fill_type='solid'),
    'target': PatternFill(start_color='E2F3E5', end_color='E2F3E5', fill_type='solid'),
    'exceed': PatternFill(start_color='E9E2F7', end_color='E9E2F7', fill_type='solid'),
}

# (key, label, kind) -- same set/order as app.py's render_metric_cards(), kind picks
# the cell's Excel number_format so rates render as %, money as thousands-separated,
# etc. (matching how the Streamlit metric cards display them).
_METRIC_ROWS = [
    ('total_orders', 'Total Orders', 'count'),
    ('delivered_rate', 'Delivered rate', 'pct'),
    ('returned_rate', 'Returned rate', 'pct'),
    ('cancelled_rate', 'Cancelled rate', 'pct'),
    ('pending_rate', 'Pending rate', 'pct'),
    ('fulfillment_lead_time_days', 'Avg. fulfillment lead time (days)', 'days'),
    ('delivery_time_days', 'Avg. delivery time (days)', 'days'),
    ('on_time_rate', 'On-time delivery rate', 'pct'),
    ('net_delivery_rate', 'Net delivery rate', 'pct'),
    ('new_rate', 'New-customer rate', 'pct'),
    ('returning_rate', 'Returning-customer rate', 'pct'),
    ('total_order_value', 'Total order value', 'money'),
    ('avg_order_value', 'Avg. order value', 'money'),
]


# Per-country table columns -- '__count_<Status>__' / '__value_<Status>__' are synthetic
# keys resolved via _metric_value() below (pulled from metrics['status_counts'] /
# metrics['status_value'], not a flat metrics.get() lookup like every other row here) so
# both the raw order COUNT and the money behind it sit right next to each status's rate
# (Sep 2026, per Mahmoud: first "الوصل بكام و الرجع بكام و الاتكنسل بكام و البيندينج
# بكام" -- the counts, not just the % -- then "عاوز احسب فلوس الاوردرات الوصلت و الرجعت
# و الاتكنسل والبندج" -- the money behind each status too, not just how many orders).
_PER_COUNTRY_METRICS = [
    ('total_orders', 'Total Orders', 'count'),
    ('__count_Delivered__', 'Delivered (count)', 'count'),
    ('__value_Delivered__', 'Delivered (value)', 'money'),
    ('delivered_rate', 'Delivered rate', 'pct'),
    ('__count_Returned__', 'Returned (count)', 'count'),
    ('__value_Returned__', 'Returned (value)', 'money'),
    ('returned_rate', 'Returned rate', 'pct'),
    ('__count_Cancelled__', 'Cancelled (count)', 'count'),
    ('__value_Cancelled__', 'Cancelled (value)', 'money'),
    ('cancelled_rate', 'Cancelled rate', 'pct'),
    ('__count_Pending__', 'Pending (count)', 'count'),
    ('__value_Pending__', 'Pending (value)', 'money'),
    ('pending_rate', 'Pending rate', 'pct'),
    ('fulfillment_lead_time_days', 'Fulfillment lead time (d)', 'days'),
    ('delivery_time_days', 'Delivery time (d)', 'days'),
    ('on_time_rate', 'On-time rate', 'pct'),
    ('net_delivery_rate', 'Net delivery rate', 'pct'),
    ('new_rate', 'New-customer rate', 'pct'),
    ('returning_rate', 'Returning-customer rate', 'pct'),
    ('total_order_value', 'Total order value', 'money'),
]

# (rate_key, label, status_name) -- status_name indexes metrics['status_counts'] /
# metrics['status_value'] for the raw count/money columns in the Comparison sheet's
# per-country tables.
_COMPARISON_METRICS = [
    ('delivered_rate', 'Delivered', 'Delivered'),
    ('cancelled_rate', 'Cancelled', 'Cancelled'),
    ('returned_rate', 'Returned', 'Returned'),
    ('pending_rate', 'Pending', 'Pending'),
]

_NUMBER_FORMATS = {'pct': '0.0%', 'days': '0.0"d"', 'money': '#,##0', 'count': '#,##0'}


def _metric_value(m, key):
    """Resolves a _PER_COUNTRY_METRICS/_METRIC_ROWS key against a metrics dict -- a
    '__count_<Status>__' key pulls from the nested status_counts dict, a
    '__value_<Status>__' key from the nested status_value dict, instead of a flat
    metrics.get(key) lookup like every other row."""
    if key.startswith('__count_'):
        return m.get('status_counts', {}).get(key[len('__count_'):-2])
    if key.startswith('__value_'):
        return m.get('status_value', {}).get(key[len('__value_'):-2])
    return m.get(key)


def _period_label(start_date, end_date):
    """Human label for a period -- the month name/year if the range starts on the 1st
    and doesn't cross into another calendar month (Sep 2026, per Mahmoud: "اسامي
    الشهور بدل period a و b"), otherwise the plain date range. Deliberately NOT
    requiring the end date to land on the exact last day of the month -- the sidebar's
    date picker is bounded by the data actually loaded (see app.py), so "all of August"
    often comes out as Aug 1-30 rather than Aug 1-31, and that should still read as
    "Aug 2026", not fall back to a date range."""
    start = pd.Timestamp(start_date)
    end = pd.Timestamp(end_date)
    if start.day == 1 and end.year == start.year and end.month == start.month:
        return start.strftime('%b %Y')
    return f"{start.date()} to {end.date()}"


def _format_on_time_target_meta(on_time_target_days):
    """meta['on_time_target_days'] is now a per-market dict (see app.py's sidebar) --
    renders it as 'UAE 3d, OM 10d, SA 7d, ...' for the workbook's title block instead of
    the single flat number this used to be."""
    if on_time_target_days is None:
        return "not set"
    if isinstance(on_time_target_days, dict):
        return ', '.join(f"{c} {d}d" for c, d in sorted(on_time_target_days.items()) if d is not None) or "not set"
    return f"{on_time_target_days} day(s)"


def _xlsx_bytes(wb):
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def _write_title_block(ws, lines, start_row=1):
    r = start_row
    for i, line in enumerate(lines):
        cell = ws.cell(row=r, column=1, value=line)
        cell.font = _XLSX_TITLE_FONT if i == 0 else _XLSX_SUBTITLE_FONT
        r += 1
    return r + 1  # first free row, with a blank-row gap


def _write_header_row(ws, row, left_col, headers):
    for c, name in enumerate(headers, start=left_col):
        cell = ws.cell(row=row, column=c, value=name)
        cell.font = _XLSX_HEADER_FONT
        cell.fill = _XLSX_HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        letter = get_column_letter(c)
        # max(), never a blind overwrite -- a column two tables share (e.g. Overview's
        # column A, used by both the metric table and the status-count table below it)
        # keeps whichever width fits its widest content instead of a later, narrower
        # table shrinking it back down and truncating long labels above it.
        wanted = max(16, len(str(name)) + 2)
        current = ws.column_dimensions[letter].width
        ws.column_dimensions[letter].width = max(wanted, current) if current else wanted


def _write_band_cell(ws, row, col, band):
    """Writes one Band cell (BAND_LABELS text + matching _BAND_FILLS colour), or '—' with
    no fill if band is None (metric not bandable here -- e.g. no window configured)."""
    cell = ws.cell(row=row, column=col, value=BAND_LABELS.get(band, '—'))
    cell.font = _XLSX_BODY_FONT
    if band in _BAND_FILLS:
        cell.fill = _BAND_FILLS[band]
    return cell


def _write_metric_value_table(ws, top_row, left_col, metric_defs, metrics, band_fn=None):
    """Metric | Value [| Band] table -- one row per (key, label, kind) in metric_defs,
    value pulled from metrics[key]. band_fn(key, value) -> 'below'/'target'/'exceed'/None
    (Sep 2026) adds a 3rd Band column, colour-filled to match the on-screen captions --
    omit it for metrics with no CEO-scorecard band (most of them). Returns
    (first_data_row, last_data_row)."""
    headers = ['Metric', 'Value'] + (['Band'] if band_fn else [])
    _write_header_row(ws, top_row, left_col, headers)
    r = top_row + 1
    for key, label, kind in metric_defs:
        ws.cell(row=r, column=left_col, value=label).font = _XLSX_BODY_FONT
        val_cell = ws.cell(row=r, column=left_col + 1, value=metrics.get(key))
        val_cell.font = _XLSX_BODY_FONT
        val_cell.number_format = _NUMBER_FORMATS[kind]
        if band_fn:
            _write_band_cell(ws, r, left_col + 2, band_fn(key, metrics.get(key)))
        r += 1
    ws.column_dimensions[get_column_letter(left_col)].width = 36
    return top_row + 1, r - 1


def _write_comparison_metric_table(ws, top_row, left_col, metrics_a, metrics_b, deltas, label_a, label_b, band_fn=None):
    """Metric | <label A> [| A band] | <label B> [| B band] | Delta | Unit [| Trend] --
    same rows as app.py's Comparison tab 'Overall deltas' table. label_a/label_b: human
    period labels (see _period_label()) shown instead of the generic 'Period A'/'Period
    B'. band_fn(key, value) -> band or None (Sep 2026): when given, adds a Band column
    next to EACH period's value (so a metric can read "Below target" for A and "On
    target" for B, or vice versa) plus a Trend column (Improving/Declining/Flat, from
    the SAME delta/direction logic as the Delta column -- trend and target-achievement
    are 2 separate questions, answered side by side rather than one standing in for the
    other; see app.py's Comparison tab caption for the same idea on screen)."""
    headers = ['Metric', label_a] + ([f'{label_a} band'] if band_fn else []) \
        + [label_b] + ([f'{label_b} band'] if band_fn else []) + ['Delta', 'Unit'] \
        + (['Trend'] if band_fn else [])
    _write_header_row(ws, top_row, left_col, headers)
    r = top_row + 1
    for key, label in METRIC_LABELS.items():
        d = deltas.get(key)
        if d is None:
            continue
        kind = 'pct' if key.endswith('_rate') else 'days'
        unit = 'pp' if kind == 'pct' else 'days'
        va, vb = metrics_a.get(key), metrics_b.get(key)
        ws.cell(row=r, column=left_col, value=label).font = _XLSX_BODY_FONT
        col = left_col + 1
        for val in (va, vb):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = _XLSX_BODY_FONT
            cell.number_format = _NUMBER_FORMATS[kind]
            col += 1
            if band_fn:
                _write_band_cell(ws, r, col, band_fn(key, val))
                col += 1
        delta_cell = ws.cell(row=r, column=col, value=d / 100 if kind == 'pct' else d)
        delta_cell.font = _XLSX_BODY_FONT
        delta_cell.number_format = '+0.0%;-0.0%' if kind == 'pct' else '+0.0"d";-0.0"d"'
        col += 1
        ws.cell(row=r, column=col, value=unit).font = _XLSX_BODY_FONT
        col += 1
        if band_fn:
            direction = METRIC_DIRECTION.get(key, 1)
            trend = ('Improving' if d * direction > 0 else 'Declining') if d != 0 else 'Flat'
            ws.cell(row=r, column=col, value=trend).font = _XLSX_BODY_FONT
        r += 1
    ws.column_dimensions[get_column_letter(left_col)].width = 34
    return top_row + 1, r - 1


def _write_status_table_and_chart(ws, top_row, left_col, metrics, title):
    """Status | Count | Value table, with its bar chart (count-based) anchored BELOW
    the table (not beside it) -- keeps the chart clear of whatever table sits to its
    right (e.g. this function is called twice per row for Period A/B side by side;
    anchoring below instead of beside removes any risk of the two charts, or a chart
    and a neighbouring table, overlapping regardless of exact column widths). The Value
    column is the money behind each status (Sep 2026, per Mahmoud: "عاوز احسب فلوس
    الاوردرات الوصلت و الرجعت و الاتكنسل والبندج")."""
    headers = ['Status', 'Count', 'Value']
    _write_header_row(ws, top_row, left_col, headers)
    r = top_row + 1
    for s in STATUSES:
        ws.cell(row=r, column=left_col, value=s).font = _XLSX_BODY_FONT
        cell = ws.cell(row=r, column=left_col + 1, value=metrics['status_counts'][s])
        cell.font = _XLSX_BODY_FONT
        cell.number_format = _NUMBER_FORMATS['count']
        value_cell = ws.cell(row=r, column=left_col + 2, value=metrics['status_value'][s])
        value_cell.font = _XLSX_BODY_FONT
        value_cell.number_format = _NUMBER_FORMATS['money']
        r += 1
    last_row = r - 1

    chart = BarChart()
    chart.type = 'bar'
    chart.title = title
    chart.legend = None
    chart.height, chart.width = 7, 12
    data = Reference(ws, min_col=left_col + 1, min_row=top_row, max_row=last_row)
    cats = Reference(ws, min_col=left_col, min_row=top_row + 1, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # No per-bar data labels here (Sep 2026, per Mahmoud: they rendered as an
    # unreadable jumble of "<series>, <status>, <value>" text once opened in Google
    # Sheets, which doesn't honour the individual show-flags that keep them clean in
    # real Excel/LibreOffice -- see the module-level note by the imports). The status
    # name is still the x-axis tick label, and the exact count/value is one column over
    # in the table right next to the chart.
    chart_row = last_row + 2
    ws.add_chart(chart, f"{get_column_letter(left_col)}{chart_row}")
    return chart_row + 14  # bottom row of the chart, roughly (14 rows tall @ ~7cm)


def _write_per_country_sheet(ws, per_country, title):
    ws.cell(row=1, column=1, value=title).font = _XLSX_TITLE_FONT
    top_row = 3
    headers = ['Country'] + [label for _, label, _ in _PER_COUNTRY_METRICS]
    _write_header_row(ws, top_row, 1, headers)
    r = top_row + 1
    for country in sorted(per_country):
        m = per_country[country]
        ws.cell(row=r, column=1, value=country).font = _XLSX_BODY_FONT
        for c, (key, _, kind) in enumerate(_PER_COUNTRY_METRICS, start=2):
            cell = ws.cell(row=r, column=c, value=_metric_value(m, key))
            cell.font = _XLSX_BODY_FONT
            cell.number_format = _NUMBER_FORMATS[kind]
        r += 1
    last_row = r - 1
    if last_row < top_row + 1:
        return  # no countries in this slice -- nothing to chart

    col_of = {key: c for c, (key, _, _) in enumerate(_PER_COUNTRY_METRICS, start=2)}
    chart_row = last_row + 3
    anchor_col = 1
    for key, label in [('delivered_rate', 'Delivered rate'), ('cancelled_rate', 'Cancelled rate'),
                        ('returned_rate', 'Returned rate')]:
        chart = BarChart()
        chart.type = 'col'
        chart.title = f"{label} by country"
        chart.legend = None
        chart.y_axis.numFmt = '0%'
        chart.x_axis.title = 'Country'
        chart.height, chart.width = 7, 11
        data = Reference(ws, min_col=col_of[key], min_row=top_row, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=top_row + 1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        # No per-bar data labels (see the note in _write_status_table_and_chart above)
        # -- country is the x-axis tick label, exact numbers are in the table above.
        ws.add_chart(chart, f"{get_column_letter(anchor_col)}{chart_row}")
        anchor_col += 8


def _write_kpi_band_table(ws, top_row, left_col, rows, title):
    """Writes per_country_kpi_rows()'s output as one table: Country | Net delivery rate
    | ND band | On-time rate | OT band | Delivery time (d) | Window | DT band -- the
    xlsx twin of app.py's on-screen "Logistics KPIs by market" table (Sep 2026), same
    columns, same band colours. Returns the next free row (for stacking a 2nd table,
    e.g. Period A then Period B, below this one)."""
    ws.cell(row=top_row, column=left_col, value=title).font = _XLSX_SUBTITLE_FONT
    header_row = top_row + 1
    _write_header_row(ws, header_row, left_col, [
        'Country', 'Net delivery rate', 'ND band', 'On-time rate', 'OT band',
        'Delivery time (d)', 'Window', 'DT band',
    ])
    r = header_row + 1
    for row in rows:
        window = row['delivery_window']
        ws.cell(row=r, column=left_col, value=row['country']).font = _XLSX_BODY_FONT

        nd_cell = ws.cell(row=r, column=left_col + 1)
        nd_cell.font = _XLSX_BODY_FONT
        if row['net_delivery_matured']:
            nd_cell.value = row['net_delivery_rate']
            nd_cell.number_format = _NUMBER_FORMATS['pct']
        elif row.get('shipped_n'):
            nd_cell.value = f"{(row.get('in_transit_share') or 0) * 100:.0f}% in transit"
        _write_band_cell(ws, r, left_col + 2, row['net_delivery_band'])

        ot_cell = ws.cell(row=r, column=left_col + 3, value=row['on_time_rate'])
        ot_cell.font = _XLSX_BODY_FONT
        ot_cell.number_format = _NUMBER_FORMATS['pct']
        _write_band_cell(ws, r, left_col + 4, row['on_time_band'])

        dt_cell = ws.cell(row=r, column=left_col + 5, value=row['delivery_time_days'])
        dt_cell.font = _XLSX_BODY_FONT
        dt_cell.number_format = _NUMBER_FORMATS['days']
        ws.cell(row=r, column=left_col + 6,
                value=(f"{window[0]}–{window[1]}d" if window else '—')).font = _XLSX_BODY_FONT
        _write_band_cell(ws, r, left_col + 7, row['delivery_time_band'])
        r += 1

    for i in range(8):
        letter = get_column_letter(left_col + i)
        wanted = 18
        current = ws.column_dimensions[letter].width
        ws.column_dimensions[letter].width = max(wanted, current) if current else wanted
    return r + 1


def _band_fn_from_meta(meta):
    """Builds a band_fn(key, value) -> band closure from meta's 'delivery_windows',
    'on_time_bands', 'net_delivery_bands' (Sep 2026 -- see app.py's export_meta). Any of
    the 3 missing from meta just means that metric won't be banded (returns None for
    it) -- lets both export functions stay callable even if a caller doesn't pass the
    new keys. delivery_time_days only bands when exactly one country is selected (its
    window varies by market -- see app.py's _single_country for the same rule on
    screen)."""
    delivery_windows = meta.get('delivery_windows') or {}
    on_time_bands = meta.get('on_time_bands')
    net_delivery_bands = meta.get('net_delivery_bands')
    countries = meta.get('countries') or []
    single_country = countries[0] if len(countries) == 1 else None

    def band_fn(key, value):
        if key == 'on_time_rate' and on_time_bands:
            return classify_band(value, *on_time_bands)
        if key == 'net_delivery_rate' and net_delivery_bands:
            return classify_band(value, *net_delivery_bands)
        if key == 'delivery_time_days' and single_country in delivery_windows:
            return classify_delivery_time_band(value, delivery_windows[single_country])
        return None
    return band_fn


def export_single_period_xlsx(metrics, meta):
    """metrics: compute_period_metrics()'s return value. meta: dict with 'countries'
    (list), 'on_time_target_days', 'generated_at' (str), plus (Sep 2026, so the
    Below/Target/Exceed bands shown on screen make it into the download too --
    optional, omitting them just means no Band column/sheet) 'delivery_windows'
    ({country: (low, high)}), 'on_time_bands' / 'net_delivery_bands' ((target, exceed)
    tuples). Returns .xlsx bytes with the same numbers/breakdown/bands as the Overview
    tab in "Single period" mode."""
    band_fn = _band_fn_from_meta(meta)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Overview'
    r = _write_title_block(ws, [
        'Ops Pulse -- Single period report',
        f"Order Date {metrics['start_date']} → {metrics['end_date']}",
        f"Countries: {', '.join(meta['countries'])}",
        f"On-time delivery target (per market): {_format_on_time_target_meta(meta['on_time_target_days'])}",
        f"Generated: {meta['generated_at']}",
    ])
    _, last = _write_metric_value_table(ws, r, 1, _METRIC_ROWS, metrics, band_fn=band_fn)
    _write_status_table_and_chart(ws, last + 3, 1, metrics, 'Orders by status')

    ws2 = wb.create_sheet('Per Country')
    _write_per_country_sheet(ws2, metrics['per_country'], f"{metrics['start_date']} → {metrics['end_date']} -- by country")

    if meta.get('on_time_bands') and meta.get('net_delivery_bands'):
        ws3 = wb.create_sheet('Logistics KPIs')
        rows = per_country_kpi_rows(metrics, meta.get('delivery_windows') or {}, meta['on_time_bands'], meta['net_delivery_bands'])
        _write_kpi_band_table(ws3, 1, 1, rows,
                               f"{metrics['start_date']} → {metrics['end_date']} -- Net/On-time delivery vs. CEO scorecard bands")

    return _xlsx_bytes(wb)


def export_comparison_xlsx(comparison, summary, meta):
    """comparison: compare_periods()'s return value. summary: generate_summary()'s
    return value. meta: dict with 'countries', 'on_time_target_days', 'generated_at',
    plus (Sep 2026, same as export_single_period_xlsx -- optional) 'delivery_windows',
    'on_time_bands', 'net_delivery_bands'. Returns .xlsx bytes covering the Overview +
    Comparison + Summary tabs in "Compare two periods" mode, PLUS a "Logistics KPIs"
    sheet when bands are supplied -- the trend (Delta) and each period's own
    Below/Target/Exceed status shown together, the same "improving AND on/off target
    are 2 different questions" idea as the on-screen Comparison tab."""
    band_fn = _band_fn_from_meta(meta)
    metrics_a, metrics_b, deltas = comparison['period_a'], comparison['period_b'], comparison['deltas']
    label_a = _period_label(metrics_a['start_date'], metrics_a['end_date'])
    label_b = _period_label(metrics_b['start_date'], metrics_b['end_date'])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Overview'
    r = _write_title_block(ws, [
        'Ops Pulse -- Comparison report',
        f"{label_a} (baseline): {metrics_a['start_date']} → {metrics_a['end_date']}",
        f"{label_b}: {metrics_b['start_date']} → {metrics_b['end_date']}",
        f"Countries: {', '.join(meta['countries'])}",
        f"On-time delivery target (per market): {_format_on_time_target_meta(meta['on_time_target_days'])}",
        f"Generated: {meta['generated_at']}",
    ])
    _, last = _write_comparison_metric_table(ws, r, 1, metrics_a, metrics_b, deltas, label_a, label_b, band_fn=band_fn)
    chart_row = last + 3
    bottom_a = _write_status_table_and_chart(ws, chart_row, 1, metrics_a, f"{label_a} -- by status")
    _write_status_table_and_chart(ws, chart_row, 10, metrics_b, f"{label_b} -- by status")

    ws_a = wb.create_sheet('Per Country - Period A')
    _write_per_country_sheet(ws_a, metrics_a['per_country'], f"{label_a} ({metrics_a['start_date']} → {metrics_a['end_date']})")
    ws_b = wb.create_sheet('Per Country - Period B')
    _write_per_country_sheet(ws_b, metrics_b['per_country'], f"{label_b} ({metrics_b['start_date']} → {metrics_b['end_date']})")

    if meta.get('on_time_bands') and meta.get('net_delivery_bands'):
        ws_kpi = wb.create_sheet('Logistics KPIs')
        delivery_windows = meta.get('delivery_windows') or {}
        rows_a = per_country_kpi_rows(metrics_a, delivery_windows, meta['on_time_bands'], meta['net_delivery_bands'])
        rows_b = per_country_kpi_rows(metrics_b, delivery_windows, meta['on_time_bands'], meta['net_delivery_bands'])
        next_row = _write_kpi_band_table(ws_kpi, 1, 1, rows_a, f"{label_a} ({metrics_a['start_date']} → {metrics_a['end_date']})")
        _write_kpi_band_table(ws_kpi, next_row + 1, 1, rows_b, f"{label_b} ({metrics_b['start_date']} → {metrics_b['end_date']})")

    ws_cmp = wb.create_sheet('Comparison')
    ws_cmp.cell(row=1, column=1, value=f"{label_a} (baseline) vs. {label_b}, by market").font = _XLSX_TITLE_FONT
    row_cursor = 3
    for rate_key, label, status_name in _COMPARISON_METRICS:
        countries_here = sorted(set(metrics_a['per_country']) | set(metrics_b['per_country']))
        rows = []
        for c in countries_here:
            ma = metrics_a['per_country'].get(c, {})
            mb = metrics_b['per_country'].get(c, {})
            va, vb = ma.get(rate_key), mb.get(rate_key)
            if va is None and vb is None:
                continue
            ca = ma.get('status_counts', {}).get(status_name)
            cb = mb.get('status_counts', {}).get(status_name)
            # The money behind this status, per country/period (Sep 2026, per Mahmoud:
            # "عاوز احسب فلوس الاوردرات الوصلت و الرجعت و الاتكنسل والبندج").
            fa = ma.get('status_value', {}).get(status_name)
            fb = mb.get('status_value', {}).get(status_name)
            rows.append((c, ca, cb, fa, fb, va, vb))
        if not rows:
            continue
        ws_cmp.cell(row=row_cursor, column=1, value=f"{label} by country").font = _XLSX_SUBTITLE_FONT
        table_row = row_cursor + 1
        _write_header_row(ws_cmp, table_row, 1, [
            'Country', f"{label_a} (count)", f"{label_b} (count)",
            f"{label_a} (value)", f"{label_b} (value)",
            f"{label_a} (rate)", f"{label_b} (rate)",
        ])
        r2 = table_row + 1
        for c, ca, cb, fa, fb, va, vb in rows:
            ws_cmp.cell(row=r2, column=1, value=c).font = _XLSX_BODY_FONT
            for col_offset, val, kind in (
                (2, ca, 'count'), (3, cb, 'count'), (4, fa, 'money'), (5, fb, 'money'),
                (6, va, 'pct'), (7, vb, 'pct'),
            ):
                cell = ws_cmp.cell(row=r2, column=col_offset, value=val)
                cell.font = _XLSX_BODY_FONT
                cell.number_format = _NUMBER_FORMATS[kind]
            r2 += 1
        last_row2 = r2 - 1

        chart = BarChart()
        chart.type = 'col'
        chart.title = f"{label} rate by country — {label_a} vs {label_b}"
        chart.height, chart.width = 7, 12
        chart.y_axis.numFmt = '0%'
        chart.x_axis.title = 'Country'
        data = Reference(ws_cmp, min_col=6, max_col=7, min_row=table_row, max_row=last_row2)
        cats = Reference(ws_cmp, min_col=1, min_row=table_row + 1, max_row=last_row2)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        # Legend stays on here (unlike the single-series charts above) since there are
        # 2 series (label_a vs label_b) sharing each country's pair of bars -- the
        # legend is what tells them apart. No per-bar data labels (see the note in
        # _write_status_table_and_chart) -- the exact count/value/rate are all right
        # there in the table above the chart.
        ws_cmp.add_chart(chart, f"J{row_cursor}")

        row_cursor = max(last_row2, row_cursor + 15) + 3

    ws_sum = wb.create_sheet('Summary')
    ws_sum.cell(row=1, column=1, value='✅ What\'s working').font = _XLSX_TITLE_FONT
    r = 3
    _write_header_row(ws_sum, r, 1, ['Scope', 'Metric', 'Message'])
    r += 1
    if summary['good']:
        for item in summary['good']:
            ws_sum.cell(row=r, column=1, value=item['scope']).font = _XLSX_BODY_FONT
            ws_sum.cell(row=r, column=2, value=METRIC_LABELS.get(item['metric'], item['metric'])).font = _XLSX_BODY_FONT
            ws_sum.cell(row=r, column=3, value=item['message']).font = _XLSX_BODY_FONT
            r += 1
    else:
        ws_sum.cell(row=r, column=1, value='Nothing moved enough (given the thresholds used) to call it a clear improvement.').font = _XLSX_BODY_FONT
        r += 1

    r += 2
    ws_sum.cell(row=r, column=1, value='⚠️ Weak points').font = _XLSX_TITLE_FONT
    r += 2
    _write_header_row(ws_sum, r, 1, ['Scope', 'Metric', 'Message'])
    r += 1
    if summary['weak_points']:
        for item in summary['weak_points']:
            ws_sum.cell(row=r, column=1, value=item['scope']).font = _XLSX_BODY_FONT
            ws_sum.cell(row=r, column=2, value=METRIC_LABELS.get(item['metric'], item['metric'])).font = _XLSX_BODY_FONT
            ws_sum.cell(row=r, column=3, value=item['message']).font = _XLSX_BODY_FONT
            r += 1
    else:
        ws_sum.cell(row=r, column=1, value='No weak points crossed the thresholds in this comparison.').font = _XLSX_BODY_FONT
        r += 1

    r += 2
    ws_sum.cell(row=r, column=1, value='Thresholds used').font = _XLSX_TITLE_FONT
    r += 2
    _write_header_row(ws_sum, r, 1, ['Threshold', 'Value'])
    r += 1
    for key, val in summary['thresholds_used'].items():
        ws_sum.cell(row=r, column=1, value=key).font = _XLSX_BODY_FONT
        ws_sum.cell(row=r, column=2, value=val).font = _XLSX_BODY_FONT
        r += 1
    ws_sum.column_dimensions['A'].width = 14
    ws_sum.column_dimensions['B'].width = 30
    ws_sum.column_dimensions['C'].width = 80

    return _xlsx_bytes(wb)
